#!/usr/bin/env python3
import asyncio
import os
from pathlib import Path
from typing import List, Dict

from src.milvus_connector import get_embedding
from constants import COLLECTION_NAME

try:
    from docx import Document  # python-docx
except ImportError:
    Document = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    from pypdf import PdfReader
except ImportError:
    PdfReader = None

def read_text_from_file(file_path: Path) -> str:
    suffix = file_path.suffix.lower()

    if suffix in {".txt", ".md"}:
        return file_path.read_text(encoding="utf-8", errors="ignore")

    if suffix == ".docx":
        if Document is None:
            raise RuntimeError("python-docx is required to read .docx files. Please install it.")
        doc = Document(str(file_path))
        paragraphs = [p.text for p in doc.paragraphs if p.text and p.text.strip()]
        return "\n".join(paragraphs)

    if suffix in {".xlsx", ".xlsm"}:
        if load_workbook is None:
            raise RuntimeError("openpyxl is required to read .xlsx files. Please install it.")
        wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
        parts: List[str] = []
        for ws in wb.worksheets:
            parts.append(f"# Sheet: {ws.title}")
            for row in ws.iter_rows(values_only=True):
                row_text = " \t ".join([str(c) if c is not None else "" for c in row])
                if row_text.strip():
                    parts.append(row_text)
        wb.close()
        return "\n".join(parts)

    if suffix == ".pdf":
        if PdfReader is None:
            raise RuntimeError("pypdf is required to read .pdf files. Please install it.")
        try:
            reader = PdfReader(str(file_path))
            pages_text: List[str] = []
            for page in reader.pages:
                extracted = page.extract_text() or ""
                if extracted.strip():
                    pages_text.append(extracted)
            return "\n".join(pages_text)
        except Exception:
            # If extraction fails (e.g., scanned PDFs), return empty string so caller can skip
            return ""

    # Unsupported file types are skipped gracefully
    return ""


def chunk_text(text: str, chunk_size: int = 1200, chunk_overlap: int = 200) -> List[str]:
    if not text:
        return []
    words = text.split()
    chunks: List[str] = []
    start = 0
    while start < len(words):
        end = min(start + chunk_size, len(words))
        chunks.append(" ".join(words[start:end]))
        if end == len(words):
            break
        start = max(end - chunk_overlap, start + 1)
    return chunks


def get_max_existing_chunk_id(output_dir: Path) -> int:
    import json
    if not output_dir.exists():
        return -1
    max_id = -1
    for json_path in sorted(output_dir.glob("*.json")):
        try:
            with json_path.open("r", encoding="utf-8") as f:
                data = json.load(f)
            for rec in data:
                cid = rec.get("chunk_id")
                if isinstance(cid, int) and cid > max_id:
                    max_id = cid
        except Exception:
            continue
    return max_id


async def ingest_folder(folder: Path, output_dir: Path, start_chunk_id: int = 0) -> None:
    if not folder.exists() or not folder.is_dir():
        raise FileNotFoundError(f"Folder not found: {folder}")

    supported_ext = {".txt", ".md", ".docx", ".xlsx", ".xlsm", ".pdf"}
    files: List[Path] = [p for p in folder.rglob("*") if p.is_file() and p.suffix.lower() in supported_ext]

    print(f"Discovered {len(files)} files to ingest from {folder}")

    for file_path in files:
        try:
            print(f"Reading: {file_path}")
            text = read_text_from_file(file_path)
            if not text.strip():
                print(f"Skipping empty/unsupported: {file_path}")
                continue

            chunks = chunk_text(text)
            print(f"Chunked into {len(chunks)} chunks")

            embedding_tasks = [get_embedding(chunk) for chunk in chunks]
            vectors = await asyncio.gather(*embedding_tasks, return_exceptions=True)

            records: List[Dict] = []
            doc_id = file_path.stem
            source = str(file_path)
            current_id = start_chunk_id
            for idx, (chunk, vector) in enumerate(zip(chunks, vectors)):
                if isinstance(vector, Exception):
                    vector = []
                records.append({
                    "doc_id": doc_id,
                    "chunk_id": current_id,
                    "text": chunk,
                    "source": source,
                    "vector": vector,
                })
                current_id += 1

            output_dir.mkdir(parents=True, exist_ok=True)
            out_path = output_dir / f"{file_path.stem}.json"
            import json
            payload = {"collectionName": COLLECTION_NAME, "data": records}
            with out_path.open("w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
            print(f"Wrote {len(records)} chunks to {out_path} (chunk_id {start_chunk_id}..{current_id-1})")
            start_chunk_id = current_id
        except Exception as e:
            print(f"Failed to ingest {file_path}: {e}")


async def main():
    root = Path(__file__).resolve().parents[1]
    folder = root / "milvusfiles"
    output_dir = root / "milvus_exports"
    start_id = get_max_existing_chunk_id(output_dir) + 1
    await ingest_folder(folder, output_dir, start_chunk_id=start_id)


if __name__ == "__main__":
    asyncio.run(main())


